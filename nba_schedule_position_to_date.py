from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re
import pandas as pd

def load_workbook_range(range_string, ws):
    '''
    :param range_string: string of cells
    :param ws: worksheet in excel workbook
    :return: df of those cells
    '''
    col_start, col_end = re.findall("[A-Z]+", range_string)
    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))


workbook = load_workbook(filename = r"C:\Users\rober\OneDrive\Documents\Roberts Side Projects\May 22 output\eastern_conf_LP_model_may_29.xlsx")
ws = workbook.active
three_game_trips = load_workbook_range('C93:Q107', ws)
three_game_bool = load_workbook_range('U111:AOD125', ws)
two_game_trips = load_workbook_range('C21:Q35', ws)
two_game_bool = load_workbook_range('U39:AOS53', ws)
four_game_trips = load_workbook_range('C57:Q71', ws)
four_game_bool = load_workbook_range('U75:ANO89', ws)
six_game_trips = load_workbook_range('C129:Q143', ws)
six_game_bool = load_workbook_range('U147:AMK161', ws)
one_game_trips = load_workbook_range('C165:Q179', ws)
one_game_bool = load_workbook_range('U183:APH197', ws)
team_number_lookup = load_workbook_range('T20:U35', ws)

home_team_number_lookup = load_workbook_range('CS19:CT34', ws)
schedule_df = load_workbook_range('DB20:HQ35', ws)


def replace_headers(df):
    new_header = df.iloc[0]
    df = df[1:]
    df.columns = new_header

    return df


team_number_lookup = replace_headers(team_number_lookup)
home_team_number_lookup = replace_headers(home_team_number_lookup)
schedule_df = replace_headers(schedule_df)


def position_tuple_list(df):
    '''
    :param df: a dataframe
    :return: a list of tuples with the positions of a "1" in the dataframe
    '''
    position_list = [(df[col][df[col].eq(1)].index[i], df.columns.get_loc(col)) for col in
                     df.columns for i in range(len(df[col][df[col].eq(1)].index))]

    return position_list


def unpack_road_trip_string(position_of_game, time_position, road_trip_string, schedule_df, trip_count_per_position):
    away_team_num = position_of_game[trip_count_per_position][0]
    print("inside unpack road trip string")
    print(away_team_num)
    away_team = team_number_lookup.loc[team_number_lookup.number == away_team_num, 'team'].values[0]

    # removing final comma from string
    road_trip_string = str(road_trip_string)
    if len(road_trip_string) > 1:
        if road_trip_string[-1] == "," :
            road_trip_string = road_trip_string[:-1]

    road_trip_list = road_trip_string.split(",")
    road_trip_w_team_name = []
    team_count = 0
    for team in road_trip_list:
        team = int(team)
        print(team)
        print(team_number_lookup.loc[home_team_number_lookup.number == team, 'team'].values[0])
        # time_position + team_count allows teams to be separated
        schedule_df.iloc[away_team_num, (time_position + team_count)*2] = away_team + "@" + team_number_lookup.loc[home_team_number_lookup.number == team, 'team'].values[0]
        team_count += 1

    road_trip_to_print = str(road_trip_w_team_name)
    road_trip_to_print = road_trip_to_print.strip("[]")
    print(road_trip_string)
    print(road_trip_to_print)

    return schedule_df



def retrieve_road_trips_from_boolean_df(boolean_df, trip_df, schedule_df):
    num_of_cols = len(boolean_df.columns)
    end_col = 15
    start_col = 0
    position = 1
    # loop goes through one set of road trips (doing three game road trips for initial coding) (should make into function)
    while end_col < num_of_cols:

        position_df = boolean_df.iloc[0:15, start_col:end_col]
        total_per_position = sum(list(position_df.sum()))
        if total_per_position > 0:

            position_of_game = position_tuple_list(position_df)
            print(position_of_game)
            trip_count_per_position = 0
            while trip_count_per_position < total_per_position:
                road_trip_string = trip_df.iloc[position_of_game[trip_count_per_position][0], position_of_game[trip_count_per_position][1]]
                print(road_trip_string)
                # calling function to unpack string of road trips
                schedule_df = unpack_road_trip_string(position_of_game, position, road_trip_string, schedule_df, trip_count_per_position)
                trip_count_per_position += 1

        start_col += 15
        end_col += 15
        position += 1

    return schedule_df


schedule_df = retrieve_road_trips_from_boolean_df(three_game_bool, three_game_trips, schedule_df)
schedule_df = retrieve_road_trips_from_boolean_df(six_game_bool, six_game_trips, schedule_df)
schedule_df = retrieve_road_trips_from_boolean_df(one_game_bool, one_game_trips, schedule_df)
schedule_df = retrieve_road_trips_from_boolean_df(two_game_bool, two_game_trips, schedule_df)
schedule_df = retrieve_road_trips_from_boolean_df(four_game_bool, four_game_trips, schedule_df)
schedule_df.to_csv(r'C:\Users\rober\OneDrive\Documents\Roberts Side Projects\May 22 output\eastern_conference_2021_jun_1_output.csv')


